"""日本外務省制裁リスト パーサー
Ministry of Foreign Affairs of Japan / Ministry of Finance
HTML/Excel形式
URL: https://www.mofa.go.jp/mofaj/gaiko/gaikoku/seisai_list.html
MOF: https://www.mof.go.jp/policy/international_policy/gaitame_kawase/gaitame/economic_sanctions/list.html
Fallback: OpenSanctions mirror of Japan MOF sanctions XLSX
"""
import csv
import io
import re
import tempfile
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from typing import Dict, Iterator, List, Optional
from .base import SanctionEntry, BaseParser

MOFA_START_URL = "https://www.mofa.go.jp/mofaj/gaiko/gaikoku/seisai_list.html"
MOFA_BASE_URL = "https://www.mofa.go.jp"

# Ministry of Finance also publishes the sanctions list
MOF_LIST_URL = "https://www.mof.go.jp/policy/international_policy/gaitame_kawase/gaitame/economic_sanctions/list.html"
MOF_BASE_URL = "https://www.mof.go.jp"

# OpenSanctions maintains a daily-updated mirror of the Japan MOF source XLSX
OPENSANCTIONS_XLSX = (
    "https://data.opensanctions.org/datasets/latest/jp_mof_sanctions/source.xlsx"
)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en;q=0.9",
}


class MOFAJapanParser(BaseParser):
    source = "mofa_japan"

    def fetch_and_parse(self) -> Iterator[SanctionEntry]:
        print("Fetching Japan MOFA/MOF Sanctions List...")

        # Strategy 1: Try MOFA page for Excel/CSV links
        file_urls = self._find_file_links(MOFA_START_URL, MOFA_BASE_URL)

        # Strategy 2: Try MOF page for Excel/CSV links
        if not file_urls:
            file_urls = self._find_file_links(MOF_LIST_URL, MOF_BASE_URL)

        # Strategy 3: Try to parse files found on either page
        # Prefer XLSX over CSV (they contain the same data)
        if file_urls:
            xlsx_urls = [u for u in file_urls if u.lower().endswith((".xlsx", ".xls"))]
            csv_urls = [u for u in file_urls if u.lower().endswith(".csv")]

            parsed = False
            for url in xlsx_urls:
                try:
                    yield from self._parse_excel_file(url)
                    parsed = True
                    break  # Only parse first successful XLSX
                except Exception as e:
                    print(f"  MOFA: Error processing {url}: {e}")

            # Only try CSV if XLSX failed
            if not parsed:
                for url in csv_urls:
                    try:
                        yield from self._parse_csv_file(url)
                        parsed = True
                        break
                    except Exception as e:
                        print(f"  MOFA: Error processing {url}: {e}")

            if parsed:
                return

        # Strategy 4: Try OpenSanctions mirror XLSX
        print("  MOFA: No direct files found, trying OpenSanctions mirror...")
        try:
            yield from self._parse_excel_file(OPENSANCTIONS_XLSX)
            return
        except Exception as e:
            print(f"  MOFA: OpenSanctions mirror failed: {e}")

        # Strategy 5: Try HTML table parsing as last resort
        print("  MOFA: Falling back to HTML table parsing")
        yield from self._parse_html_tables()

    def _find_file_links(self, page_url: str, base_url: str) -> List[str]:
        """ページからExcel/CSVリンクを抽出"""
        urls: List[str] = []
        try:
            resp = requests.get(page_url, timeout=30, headers=HEADERS)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.content, "html.parser")

            for link in soup.find_all("a", href=True):
                href = link["href"]
                if re.search(r"\.(xlsx?|csv)$", href, re.IGNORECASE):
                    if href.startswith("/"):
                        full_url = f"{base_url}{href}"
                    elif href.startswith("http"):
                        full_url = href
                    else:
                        # Relative to current page directory
                        base_dir = page_url.rsplit("/", 1)[0]
                        full_url = f"{base_dir}/{href}"
                    if full_url not in urls:
                        urls.append(full_url)
                        print(f"  MOFA: Found file link: {full_url}")

        except Exception as e:
            print(f"  MOFA: Failed to fetch page {page_url}: {e}")

        return urls

    def _parse_excel_file(self, url: str) -> Iterator[SanctionEntry]:
        """Excelファイルをダウンロードしてパース"""
        print(f"  MOFA: Downloading Excel from {url}")
        resp = requests.get(url, timeout=120, headers=HEADERS)
        resp.raise_for_status()

        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        try:
            with os.fdopen(tmp_fd, "wb") as f:
                f.write(resp.content)

            wb = load_workbook(tmp_path, read_only=True, data_only=True)

            # Skip first worksheet if it's a TOC (一覧 = "table of contents")
            toc_names = ["一覧", "目次", "toc", "index", "contents"]
            for ws in wb.worksheets:
                ws_title_lower = ws.title.strip().lower()
                if ws_title_lower in toc_names:
                    print(f"  MOFA: Skipping TOC worksheet '{ws.title}'")
                    continue
                yield from self._parse_worksheet(ws)

            wb.close()
        finally:
            os.unlink(tmp_path)

    def _parse_worksheet(self, ws) -> Iterator[SanctionEntry]:
        """ワークシートから制裁エントリを抽出"""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return

        # Detect default entity type from worksheet title
        ws_title = ws.title or ""
        if "個人" in ws_title or "individual" in ws_title.lower():
            default_entity_type = "individual"
        elif "団体" in ws_title or "銀行" in ws_title or "entity" in ws_title.lower():
            default_entity_type = "entity"
        else:
            default_entity_type = "entity"

        # Derive program name from worksheet title
        ws_program = ws_title.strip()

        # Find header row by looking for name-related keywords
        header_idx = 0
        headers: List[str] = []
        name_keywords = ["name", "名前", "氏名", "名称", "表記",
                         "entity", "告示"]

        for idx, row in enumerate(rows):
            cells = [str(c).strip().lower() if c else "" for c in row]
            if any(kw in cell for cell in cells for kw in name_keywords):
                headers = cells
                header_idx = idx
                break

        if not headers:
            # No header found, try treating each row as data
            # Assume first column is name
            for row in rows:
                if not row or all(c is None for c in row):
                    continue
                name = str(row[0]).strip() if row[0] else ""
                if not name:
                    continue
                yield SanctionEntry(
                    source="mofa_japan",
                    source_id=None,
                    entity_type="entity",
                    name_primary=name,
                    names_aliases=self._extract_alt_names(row),
                    country=str(row[1]).strip() if len(row) > 1 and row[1] else None,
                    address=None,
                    programs=["MOFA_JAPAN_SANCTIONS"],
                    reason=None,
                )
            return

        # Map columns
        col_map = {h: i for i, h in enumerate(headers) if h}

        name_jp_col = self._find_col(col_map, ["日本語表記", "名称", "氏名", "名前"])
        name_en_col = self._find_col(col_map, ["英語表記", "english", "name", "roman"])
        type_col = self._find_col(col_map, ["type", "種別", "区分", "個人・団体"])
        country_col = self._find_col(col_map, ["country", "国", "国籍", "地域"])
        alias_col = self._find_col(col_map, ["alias", "別名", "別称", "aka", "other"])
        program_col = self._find_col(col_map, ["program", "措置", "根拠", "regime", "制裁"])
        reason_col = self._find_col(col_map, ["reason", "理由", "備考", "comment"])
        id_col = self._find_col(col_map, ["id", "番号", "no", "number"])

        # Use whichever name column is available
        primary_col = name_en_col if name_en_col is not None else name_jp_col
        secondary_col = name_jp_col if primary_col == name_en_col else name_en_col

        if primary_col is None:
            # Just use the first column with content
            primary_col = 0

        count = 0
        for row in rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue

            primary_name = self._cell_str(row, primary_col)
            if not primary_name:
                continue

            # Build aliases from secondary name column and alias column
            aliases: List[str] = []
            secondary_name = self._cell_str(row, secondary_col) if secondary_col is not None else None
            if secondary_name and secondary_name != primary_name:
                aliases.append(secondary_name)

            alias_raw = self._cell_str(row, alias_col) if alias_col is not None else None
            if alias_raw:
                for part in re.split(r"[;,、]", alias_raw):
                    part = part.strip()
                    if part and part != primary_name and part not in aliases:
                        aliases.append(part)

            # Entity type: from column, or from worksheet title
            type_raw = self._cell_str(row, type_col) if type_col is not None else ""
            if type_raw and ("個人" in type_raw or "individual" in type_raw.lower()):
                entity_type = "individual"
            elif type_raw and ("団体" in type_raw or "entity" in type_raw.lower()):
                entity_type = "entity"
            else:
                entity_type = default_entity_type

            country = self._cell_str(row, country_col) if country_col is not None else None

            programs: List[str] = []
            program_raw = self._cell_str(row, program_col) if program_col is not None else None
            if program_raw:
                programs.append(program_raw)
            elif ws_program:
                programs.append(ws_program)
            else:
                programs.append("MOFA_JAPAN_SANCTIONS")

            reason = self._cell_str(row, reason_col) if reason_col is not None else None
            source_id = self._cell_str(row, id_col) if id_col is not None else None

            count += 1
            yield SanctionEntry(
                source="mofa_japan",
                source_id=source_id,
                entity_type=entity_type,
                name_primary=primary_name,
                names_aliases=aliases,
                country=country,
                address=None,
                programs=programs,
                reason=reason,
            )

        print(f"  MOFA: parsed {count} entries from worksheet '{ws.title}'")

    def _parse_csv_file(self, url: str) -> Iterator[SanctionEntry]:
        """CSVファイルをダウンロードしてパース"""
        print(f"  MOFA: Downloading CSV from {url}")
        resp = requests.get(url, timeout=120, headers=HEADERS)
        resp.raise_for_status()

        # Try multiple encodings (Japanese content may use Shift_JIS)
        text = None
        for encoding in ["utf-8-sig", "utf-8", "shift_jis", "cp932", "euc-jp"]:
            try:
                text = resp.content.decode(encoding)
                break
            except (UnicodeDecodeError, LookupError):
                continue

        if text is None:
            print(f"  MOFA CSV: Could not decode file from {url}")
            return

        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        if not rows:
            return

        # Simple parsing: first column = name, subsequent columns = metadata
        for row in rows[1:]:  # Skip header
            if not row or not row[0].strip():
                continue

            name = row[0].strip()
            aliases: List[str] = []

            # If there's a second column with an alternate name
            if len(row) > 1 and row[1].strip() and row[1].strip() != name:
                aliases.append(row[1].strip())

            country = row[2].strip() if len(row) > 2 and row[2].strip() else None

            yield SanctionEntry(
                source="mofa_japan",
                source_id=None,
                entity_type="entity",
                name_primary=name,
                names_aliases=aliases,
                country=country,
                address=None,
                programs=["MOFA_JAPAN_SANCTIONS"],
                reason=row[3].strip() if len(row) > 3 and row[3].strip() else None,
            )

    def _parse_html_tables(self) -> Iterator[SanctionEntry]:
        """HTMLページからテーブルを直接パース（ファイルが無い場合のフォールバック）"""
        for page_url in [MOFA_START_URL, MOF_LIST_URL]:
            try:
                resp = requests.get(page_url, timeout=30, headers=HEADERS)
                resp.raise_for_status()
                soup = BeautifulSoup(resp.content, "html.parser")

                found_entries = False
                for table in soup.find_all("table"):
                    rows = table.find_all("tr")
                    for row in rows[1:]:  # Skip header
                        cols = [td.get_text(strip=True) for td in row.find_all("td")]
                        if len(cols) >= 1 and cols[0]:
                            name = cols[0]
                            aliases: List[str] = []

                            # Check for English name in another column
                            if len(cols) > 1 and cols[1] and cols[1] != name:
                                aliases.append(cols[1])

                            found_entries = True
                            yield SanctionEntry(
                                source="mofa_japan",
                                source_id=None,
                                entity_type="entity",
                                name_primary=name,
                                names_aliases=aliases,
                                country=cols[2] if len(cols) > 2 else None,
                                address=None,
                                programs=["MOFA_JAPAN_SANCTIONS"],
                                reason=cols[3] if len(cols) > 3 else None,
                            )

                if found_entries:
                    return  # Found data, no need to try other URLs

            except Exception as e:
                print(f"  MOFA HTML table parsing failed for {page_url}: {e}")

    def _find_col(self, col_map: Dict[str, int], candidates: List[str]) -> Optional[int]:
        """ヘッダーマップからカラムインデックスを検索

        Prefers exact matches over substring matches to avoid
        false positives (e.g., 'type' matching 'name type').
        """
        # Pass 1: exact match
        for candidate in candidates:
            for header, idx in col_map.items():
                if header == candidate:
                    return idx

        # Pass 2: substring match
        for candidate in candidates:
            for header, idx in col_map.items():
                if candidate in header:
                    return idx

        return None

    def _cell_str(self, row: tuple, col: Optional[int]) -> Optional[str]:
        """セルの値を安全に文字列に変換"""
        if col is None or col >= len(row):
            return None
        val = row[col]
        if val is None:
            return None
        s = str(val).strip()
        return s if s and s.lower() != "none" else None

    def _extract_alt_names(self, row: tuple) -> List[str]:
        """行の残りの列から別名を抽出"""
        names: List[str] = []
        for i in range(1, len(row)):
            val = row[i]
            if val:
                s = str(val).strip()
                if s and s.lower() != "none":
                    names.append(s)
                    break  # Only take first alternate
        return names
