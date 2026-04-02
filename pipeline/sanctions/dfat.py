"""豪州DFAT制裁リスト パーサー
Department of Foreign Affairs and Trade
XLSX形式、APIキー不要
URL: https://www.dfat.gov.au/international-relations/security/sanctions/consolidated-list
Direct XLSX: https://www.dfat.gov.au/sites/default/files/regulation8_consolidated.xlsx
Fallback: OpenSanctions mirror of DFAT source XLSX
"""
import io
import re
import tempfile
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from typing import Dict, Iterator, List, Optional, Tuple
from .base import SanctionEntry, BaseParser

DFAT_PAGE_URL = "https://www.dfat.gov.au/international-relations/security/sanctions/consolidated-list"
DFAT_XLSX_DIRECT = "https://www.dfat.gov.au/sites/default/files/regulation8_consolidated.xlsx"

# OpenSanctions maintains a daily-updated mirror of the DFAT source XLSX
DFAT_OPENSANCTIONS_XLSX = (
    "https://data.opensanctions.org/datasets/latest/au_dfat_sanctions/source.xlsx"
)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
}


class DFATParser(BaseParser):
    source = "dfat"

    def fetch_and_parse(self) -> Iterator[SanctionEntry]:
        print("Fetching Australia DFAT Consolidated Sanctions List...")

        xlsx_url = self._find_xlsx_url()
        xlsx_content = self._download_xlsx(xlsx_url)

        if xlsx_content is None:
            print("  DFAT: Failed to download XLSX from all sources")
            return

        yield from self._parse_xlsx(xlsx_content)

    def _find_xlsx_url(self) -> str:
        """DFATページからXLSXダウンロードリンクを動的に探す"""
        try:
            resp = requests.get(DFAT_PAGE_URL, timeout=30, headers=HEADERS)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.content, "html.parser")

            # Look for .xlsx links on the page
            for link in soup.find_all("a", href=True):
                href = link["href"]
                if href.lower().endswith(".xlsx"):
                    # Resolve relative URLs
                    if href.startswith("/"):
                        return f"https://www.dfat.gov.au{href}"
                    elif href.startswith("http"):
                        return href

            print("  DFAT: No XLSX link found on page, using direct URL")
        except Exception as e:
            print(f"  DFAT: Page fetch failed ({e}), using direct URL")

        return DFAT_XLSX_DIRECT

    def _download_xlsx(self, url: str) -> Optional[bytes]:
        """XLSXファイルをダウンロード（複数URL順にフォールバック）"""
        # Build list of URLs to try: dynamic -> direct -> OpenSanctions mirror
        urls_to_try = [url]
        if url != DFAT_XLSX_DIRECT:
            urls_to_try.append(DFAT_XLSX_DIRECT)
        if DFAT_OPENSANCTIONS_XLSX not in urls_to_try:
            urls_to_try.append(DFAT_OPENSANCTIONS_XLSX)

        for try_url in urls_to_try:
            try:
                print(f"  Downloading XLSX from: {try_url}")
                resp = requests.get(try_url, timeout=120, headers=HEADERS)
                resp.raise_for_status()

                content_type = resp.headers.get("Content-Type", "")
                if (
                    "spreadsheet" in content_type
                    or "excel" in content_type
                    or "octet-stream" in content_type
                    or try_url.endswith(".xlsx")
                ):
                    print(f"  Downloaded {len(resp.content)} bytes")
                    return resp.content
                else:
                    print(f"  Unexpected content type: {content_type}")
            except Exception as e:
                print(f"  Download failed from {try_url}: {e}")

        return None

    def _parse_xlsx(self, content: bytes) -> Iterator[SanctionEntry]:
        """openpyxlでXLSXをパース"""
        # Write to temp file for openpyxl
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        try:
            with os.fdopen(tmp_fd, "wb") as f:
                f.write(content)

            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb.active

            if ws is None:
                print("  DFAT XLSX: No active worksheet found")
                return

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                print("  DFAT XLSX: Empty worksheet")
                return

            # Find header row (look for "Name" column)
            header_row_idx = 0
            headers_list: List[str] = []
            for idx, row in enumerate(rows):
                cells = [str(c).strip().lower() if c else "" for c in row]
                if any("name" in c for c in cells):
                    headers_list = cells
                    header_row_idx = idx
                    break

            if not headers_list:
                print(f"  DFAT XLSX: Could not find header row, first row: {rows[0]}")
                return

            print(f"  DFAT XLSX: Headers at row {header_row_idx}: {headers_list}")

            # Map column indices
            col_map = {h: i for i, h in enumerate(headers_list) if h}
            name_col = self._find_col(col_map, ["name", "full name", "name of individual or entity"])
            type_col = self._find_col(col_map, ["type", "name type", "entity type", "individual or entity"])
            country_col = self._find_col(col_map, ["country", "nationality", "citizenship"])
            address_col = self._find_col(col_map, ["address"])
            program_col = self._find_col(col_map, ["committees", "regime", "instrument of designation"])
            alias_col = self._find_col(col_map, ["alias", "other names", "also known as"])
            reason_col = self._find_col(col_map, ["reason", "comments", "additional information", "listing information"])
            ref_col = self._find_col(col_map, ["reference", "ref", "number", "id"])

            if name_col is None:
                print(f"  DFAT XLSX: Could not find name column in {list(col_map.keys())}")
                return

            # Parse data rows
            count = 0
            for row in rows[header_row_idx + 1:]:
                if not row or all(c is None for c in row):
                    continue

                name = self._cell_str(row, name_col)
                if not name:
                    continue

                # Entity type
                type_raw = self._cell_str(row, type_col) if type_col is not None else ""
                if type_raw and "individual" in type_raw.lower():
                    entity_type = "individual"
                else:
                    entity_type = "entity"

                # Aliases
                aliases: List[str] = []
                alias_raw = self._cell_str(row, alias_col) if alias_col is not None else ""
                if alias_raw:
                    # Split on semicolons or "a.k.a." patterns
                    for part in re.split(r"[;]|(?:a\.?k\.?a\.?\s*)", alias_raw, flags=re.IGNORECASE):
                        part = part.strip().strip(",").strip()
                        if part and part != name:
                            aliases.append(part)

                country = self._cell_str(row, country_col) if country_col is not None else None
                address = self._cell_str(row, address_col) if address_col is not None else None

                programs: List[str] = []
                program_raw = self._cell_str(row, program_col) if program_col is not None else ""
                if program_raw:
                    programs.append(program_raw)

                reason = self._cell_str(row, reason_col) if reason_col is not None else None
                source_id = self._cell_str(row, ref_col) if ref_col is not None else None

                count += 1
                yield SanctionEntry(
                    source="dfat",
                    source_id=source_id,
                    entity_type=entity_type,
                    name_primary=name,
                    names_aliases=aliases,
                    country=country,
                    address=address,
                    programs=programs,
                    reason=reason,
                )

            print(f"  DFAT: parsed {count} entries")
            wb.close()
        finally:
            os.unlink(tmp_path)

    def _find_col(self, col_map: Dict[str, int], candidates: List[str]) -> Optional[int]:
        """ヘッダーマップからカラムインデックスを検索

        Prefers exact matches over substring matches to avoid
        false positives (e.g., 'type' matching 'name type').
        """
        # Pass 1: exact match
        for candidate in candidates:
            for header, idx in col_map.items():
                if header == candidate:
                    return idx

        # Pass 2: substring match
        for candidate in candidates:
            for header, idx in col_map.items():
                if candidate in header:
                    return idx

        return None

    def _cell_str(self, row: tuple, col: Optional[int]) -> Optional[str]:
        """セルの値を安全に文字列に変換"""
        if col is None or col >= len(row):
            return None
        val = row[col]
        if val is None:
            return None
        s = str(val).strip()
        return s if s and s.lower() != "none" else None
