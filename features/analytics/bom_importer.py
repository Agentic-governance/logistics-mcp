"""BOM インポーター
CSV / Excel / JSON / SAP MM60 形式の BOM を BOMNode リストに変換。

入力フォーマット:
  - CSV/Excel: ヘッダー行に part_id, part_name, supplier_name,
    supplier_country, material, hs_code, quantity, unit_cost_usd, is_critical
  - JSON: [{"part_id": ..., ...}, ...]
  - SAP MM60: SAP の BOM 展開レポート (Material, Component, Plant, etc.)
"""
from __future__ import annotations

import csv
import io
import json
from typing import Optional

from features.analytics.bom_analyzer import BOMNode


class BOMImporter:
    """BOM データの入出力"""

    # CSV/Excel ヘッダーのエイリアスマッピング
    COLUMN_ALIASES = {
        "part_id": ["part_id", "id", "item_no", "item_number", "component_id"],
        "part_name": ["part_name", "name", "description", "component", "component_name", "部品名"],
        "supplier_name": ["supplier_name", "supplier", "vendor", "vendor_name", "サプライヤー"],
        "supplier_country": ["supplier_country", "country", "origin", "origin_country", "国"],
        "material": ["material", "material_type", "raw_material", "材料"],
        "hs_code": ["hs_code", "hscode", "hs", "tariff_code", "HSコード"],
        "quantity": ["quantity", "qty", "amount", "数量"],
        "unit_cost_usd": ["unit_cost_usd", "unit_cost", "cost", "price", "単価"],
        "is_critical": ["is_critical", "critical", "クリティカル"],
        "tier": ["tier", "level", "bom_level", "階層"],
    }

    def _resolve_column(self, headers: list[str], field: str) -> Optional[str]:
        """ヘッダー名のエイリアスを解決"""
        aliases = self.COLUMN_ALIASES.get(field, [field])
        for alias in aliases:
            for h in headers:
                if h.lower().strip() == alias.lower():
                    return h
        return None

    def from_csv(self, csv_text: str) -> list[BOMNode]:
        """CSV テキストから BOMNode リストに変換"""
        reader = csv.DictReader(io.StringIO(csv_text))
        headers = reader.fieldnames or []
        return self._parse_rows(list(reader), headers)

    def from_csv_file(self, filepath: str) -> list[BOMNode]:
        """CSV ファイルから BOMNode リストに変換"""
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            return self._parse_rows(list(reader), headers)

    def from_json(self, json_text: str) -> list[BOMNode]:
        """JSON テキストから BOMNode リストに変換"""
        data = json.loads(json_text)
        if isinstance(data, dict):
            # Wrapper format: {"bom": [...], "product_name": "..."}
            items = data.get("bom", data.get("parts", data.get("components", [])))
        elif isinstance(data, list):
            items = data
        else:
            return []

        return self._parse_rows(items, list(items[0].keys()) if items else [])

    def from_json_file(self, filepath: str) -> list[BOMNode]:
        """JSON ファイルから BOMNode リストに変換"""
        with open(filepath, "r", encoding="utf-8") as f:
            return self.from_json(f.read())

    def from_excel(self, filepath: str, sheet_name: str = None) -> list[BOMNode]:
        """Excel ファイルから BOMNode リストに変換"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel import. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            if not any(row):
                continue
            data.append(dict(zip(headers, row)))

        wb.close()
        return self._parse_rows(data, headers)

    def from_sap_mm60(self, csv_text: str) -> list[BOMNode]:
        """SAP MM60 (BOM 展開レポート) 形式から変換。

        SAP MM60 typical columns:
        - Level, Component, Component Description, Quantity, UoM,
          Item Category, Plant, ...

        マッピング:
        - Component → part_id
        - Component Description → part_name
        - Plant → supplier_country (SAP plant code → country lookup)
        - Quantity → quantity
        - Level → tier
        """
        # SAP plant code → country mapping (common examples)
        PLANT_COUNTRY = {
            "1000": "Germany", "1100": "Germany",
            "2000": "United States", "2100": "United States",
            "3000": "China", "3100": "China",
            "4000": "Japan", "4100": "Japan",
            "5000": "South Korea", "5100": "South Korea",
            "6000": "India",
            "7000": "Thailand",
            "8000": "Vietnam",
            "9000": "Mexico",
        }

        reader = csv.DictReader(io.StringIO(csv_text))
        nodes = []
        for row in reader:
            component = row.get("Component", row.get("Material", "")).strip()
            if not component:
                continue

            description = row.get("Component Description",
                                  row.get("Material Description", component))
            plant = row.get("Plant", "").strip()
            country = PLANT_COUNTRY.get(plant, plant)

            try:
                qty = float(row.get("Quantity", row.get("Qty", 1)))
            except (ValueError, TypeError):
                qty = 1.0

            try:
                level = int(row.get("Level", row.get("BOM Level", 1)))
            except (ValueError, TypeError):
                level = 1

            nodes.append(BOMNode(
                part_id=component,
                part_name=description,
                supplier_name=row.get("Vendor", row.get("Supplier", f"Plant {plant}")),
                supplier_country=country,
                material=row.get("Material Group", ""),
                tier=level,
                quantity=qty,
            ))

        return nodes

    def _parse_rows(self, rows: list[dict], headers: list[str]) -> list[BOMNode]:
        """汎用的な行パーサー"""
        # Resolve column names
        col_part_id = self._resolve_column(headers, "part_id")
        col_part_name = self._resolve_column(headers, "part_name")
        col_supplier = self._resolve_column(headers, "supplier_name")
        col_country = self._resolve_column(headers, "supplier_country")
        col_material = self._resolve_column(headers, "material")
        col_hs_code = self._resolve_column(headers, "hs_code")
        col_qty = self._resolve_column(headers, "quantity")
        col_cost = self._resolve_column(headers, "unit_cost_usd")
        col_critical = self._resolve_column(headers, "is_critical")
        col_tier = self._resolve_column(headers, "tier")

        nodes = []
        for i, row in enumerate(rows):
            part_id = str(row.get(col_part_id, f"P{i+1:04d}")) if col_part_id else f"P{i+1:04d}"
            part_name = str(row.get(col_part_name, "")) if col_part_name else ""
            supplier = str(row.get(col_supplier, "")) if col_supplier else ""
            country = str(row.get(col_country, "")) if col_country else ""

            if not country:
                continue

            material = str(row.get(col_material, "")) if col_material else ""
            hs_code = str(row.get(col_hs_code, "")) if col_hs_code else ""

            try:
                qty = float(row.get(col_qty, 1)) if col_qty else 1.0
            except (ValueError, TypeError):
                qty = 1.0

            try:
                cost = float(row.get(col_cost, 0)) if col_cost else 0.0
            except (ValueError, TypeError):
                cost = 0.0

            critical_val = row.get(col_critical, False) if col_critical else False
            is_critical = str(critical_val).lower() in ("true", "1", "yes", "○")

            try:
                tier = int(row.get(col_tier, 1)) if col_tier else 1
            except (ValueError, TypeError):
                tier = 1

            nodes.append(BOMNode(
                part_id=part_id,
                part_name=part_name,
                supplier_name=supplier,
                supplier_country=country,
                material=material,
                hs_code=hs_code,
                tier=tier,
                quantity=qty,
                unit_cost_usd=cost,
                is_critical=is_critical,
            ))

        return nodes

    def to_csv(self, nodes: list[BOMNode]) -> str:
        """BOMNode リストを CSV テキストに変換"""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            "part_id", "part_name", "supplier_name", "supplier_country",
            "material", "hs_code", "tier", "quantity", "unit_cost_usd", "is_critical",
        ])
        writer.writeheader()
        for n in nodes:
            writer.writerow({
                "part_id": n.part_id,
                "part_name": n.part_name,
                "supplier_name": n.supplier_name,
                "supplier_country": n.supplier_country,
                "material": n.material,
                "hs_code": n.hs_code,
                "tier": n.tier,
                "quantity": n.quantity,
                "unit_cost_usd": n.unit_cost_usd,
                "is_critical": n.is_critical,
            })
        return output.getvalue()

    def to_json(self, nodes: list[BOMNode]) -> str:
        """BOMNode リストを JSON テキストに変換"""
        data = []
        for n in nodes:
            data.append({
                "part_id": n.part_id,
                "part_name": n.part_name,
                "supplier_name": n.supplier_name,
                "supplier_country": n.supplier_country,
                "material": n.material,
                "hs_code": n.hs_code,
                "tier": n.tier,
                "quantity": n.quantity,
                "unit_cost_usd": n.unit_cost_usd,
                "is_critical": n.is_critical,
            })
        return json.dumps(data, ensure_ascii=False, indent=2)

    def to_excel(self, nodes: list[BOMNode], filepath: str):
        """BOMNode リストを Excel ファイルに出力"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"

        headers = [
            "part_id", "part_name", "supplier_name", "supplier_country",
            "material", "hs_code", "tier", "quantity", "unit_cost_usd", "is_critical",
        ]
        ws.append(headers)

        for n in nodes:
            ws.append([
                n.part_id, n.part_name, n.supplier_name, n.supplier_country,
                n.material, n.hs_code, n.tier, n.quantity, n.unit_cost_usd,
                n.is_critical,
            ])

        wb.save(filepath)
